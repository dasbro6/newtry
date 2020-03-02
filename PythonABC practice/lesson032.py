# -*- coding: gbk -*-
import openpyxl
from openpyxl.utils import get_column_letter

workbook_exp = openpyxl.Workbook()
sheet_exp = workbook_exp.active

# Change the name of sheet
print(sheet_exp.title)
sheet_exp.title  = 'expone'
print(sheet_exp.title)


workbook_exp.create_sheet(title='first_exp',index=0)
workbook_exp.create_sheet(title='second_exp',index=1)
print(workbook_exp.sheetnames)
workbook_exp.remove(workbook_exp['second_exp'])

print(workbook_exp.sheetnames)

workbook_exp.save('temp001.xlsx')

# write value in cells
fe = workbook_exp['first_exp']
eo = workbook_exp['expone']
fe['A1']='Hello Python'
print(fe['A1'].value)

for row in range(1,40):
    eo.append(range(17))

et = workbook_exp.create_sheet(title='exptwo',index=2)
rows =[['Number', 'Batch 1', 'Batch2'],
	[2, 40, 30],
	[3, 40, 25],
	[4, 50, 30],
	[5, 30, 10],
	[6, 40, 30],
	[7, 78, 52],
]

for row in rows:
    et.append(row)


eth = workbook_exp.create_sheet(title='third_exp')
for row in range(5, 30):
	for col in range(15, 54):
		eth.cell(column=col, row=row, value=get_column_letter(col))
print(eth['AA10'].value)

# workbook_exp.save(filename='empty_book.xlsx')

workbook_exp.save('temp002.xlsx')



print('End')

# Corret value in salecar.xlxs

update = {"合计":0}

workbook_sale = openpyxl.load_workbook('salecar.xlsx')
sheet1 = workbook_sale['Sheet1']

for row in range(2,sheet1.max_row+1):
	searchname = sheet1.cell(row = row,column = 4).value
	print(searchname,row)
	if searchname in update:
		for column in range(5,sheet1.max_row+1):
			print(searchname,row)
			sheet1.cell(row = row,column = column).value = update[str(searchname)]

			print(row,'complete')
workbook_sale.save('temp003.xlsx')