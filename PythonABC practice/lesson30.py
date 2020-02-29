import openpyxl

import sqlite3

salecar = openpyxl.load_workbook('salecar.xlsx')

# print(salecar.sheetnames)
#
# newsheet = salecar.create_sheet('newsheet')
#
# for sheet in salecar:
#     print(sheet.title)
#
# print(salecar.sheetnames)
#
#
# sheet3 = salecar['newsheet']
#
# print('sheet3',sheet3)
#
activesheet = salecar.active
# print('活跃表格',activesheet)
#
# print(activesheet['G10'].value)
#
# cell=activesheet['G10']
#
# print('hang {} , lie {} ,is {}'.format(cell.row,cell.column,cell.value))
# #
# print(activesheet.cell(row=7,column=10).value)
#
# # for i in range(2,30):
# #     print(activesheet.cell(row=i,column=10).value)
#
# columnG = activesheet['G']
#
# # print(columnG)
#
# column_range = activesheet['G:Z']
#
# row_range = activesheet[2:50]

# for col in column_range:
#     for cell in col:
#         print(cell,cell.value)
#
# for row in row_range:
#     for cell in row:
#         print(cell,cell.value)

# for row in activesheet.iter_rows(min_row=2,max_row=50,max_col=50):
#     for cell in row:
#         print(cell.value)

cellrange=activesheet['G2:Z55']

for rofcobj in cellrange:
    for cellobj in rofcobj:
        print(cellobj.coordinate,'\t',cellobj.value,)

    print('ending')

print('{} * {}'.format(activesheet.max_row,activesheet.max_column))

from openpyxl.utils import get_column_letter,column_index_from_string

print(get_column_letter(2),column_index_from_string('G'))